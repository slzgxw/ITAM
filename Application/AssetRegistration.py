from selenium import webdriver
# xlwt导入方式必须这样，否则闪退
import xlrd
import time
from openpyxl import load_workbook


# 需要安装的支持包有：xlrd，openpyxl，selenium


def AssetRegistraton(ExcelPath, *args, **kwargs):
    try:
        options = webdriver.ChromeOptions()

        options.add_argument(
            "user-data-dir=C:\\Users\\Administrator\\AppData\\Local\\Google\\Chrome\\User Data\\Default\\Cache")  # 携带cookice
        options.add_experimental_option("excludeSwitches", ["enable-logging"])
        wd = webdriver.Chrome(options=options, executable_path=r'./File/chromedriver.exe')

        wd.implicitly_wait(2000)
        wd.get(
            'http://asset-pre.bytedance.net/real-asset/process/asset-appropriation?current=1&pageSize=200&subStatus'
            '=1007')
        wd.implicitly_wait(2000)

        # 抓取到数据然后点击区域按钮
        element = wd.find_elements_by_xpath(
            '//*[@id="app"]/section/section/main/div[2]/div/div[1]/div[1]/div[2]/div[6]')
        for i in element:
            i.click()
        wd.implicitly_wait(2000)

        # 进入到区域input按钮自动写入地址
        element = wd.find_elements_by_xpath('//*[@id="rc-tree-select-list_1"]/span/span/input')
        for i in element:
            with open('./File/list.txt', 'r', encoding='utf-8') as file:
                content = file.read()
                i.send_keys(content)

        # 点击搜索到的地址开始查询数据
        element = wd.find_elements_by_xpath('//*[@id="rc-tree-select-list_1"]/ul/li/ul/li/ul/li/ul/li')
        for i in element:
            i.click()

        # # 点击工单来源按钮
        # time.sleep(3)  # 等待加载点击后出现地址的时间
        # element = wd.find_elements_by_xpath(
        #     '//*[@id="app"]/section/section/main/div[2]/div/div[1]/div[1]/div[2]/div[5]/div/div/span/div/div')
        # for i in element:
        #     i.click()
        #
        # time.sleep(1)  # 选择工单来源(入职领用)
        # element = wd.find_elements_by_xpath(
        #     '/html/body/div[4]/div/div/div/div/ul/li[2]')
        # for i in element:
        #     i.click()

        # 点击申请人按钮准备输入用户邮箱进行查询(完成)
        element = wd.find_elements_by_xpath(
            '//*[@id="app"]/section/section/main/div[2]/div/div[1]/div[1]/div[2]/div[3]/div')
        for i in element:
            i.click()

        element = wd.find_elements_by_xpath('/html/body/div[4]/div/div/div/div[1]/span/input')

        # pip install xlrd 需要安装此模块
        Excel = xlrd.open_workbook(r'%s' % (ExcelPath))
        table = Excel.sheets()[0]  # 获取工作表，【0】代表第一个表
        NumberOfRows = table.nrows  # 获取数据的行数
        tableList = {}  # 1：获取到的数据存入字典中，后面查询此字典内对应的IT编号
        # 2：重复的邮箱地址将只取最后一次出现的IT编号
        count = 0
        MultipleCycles = 0
        ReturnAfterRefresh = 0
        ExcelCountError = 0  # 资产信息错误填写Excel换行
        ExcelUserError = 0  # 用户信息错误填写Excel换行

        # 取Excel数据的for循环
        for x in element:
            while count <= NumberOfRows - 2:  # 拿取数据
                tableData = table.row_values(count + 1)  # 从第二行开始读取
                count += 1
                mail = tableData[1]  # 对字典key插入邮箱
                number = tableData[2]  # 对value插入IT资产
                tableList[mail] = number  # 将Excel的列转换为字典
            # 操作录入资产的for循环
            for tableLists in tableList:  # 长度不符合？ 很可能是Excel重复造成的
                returnCount = 0  # 异常判断调用，如果发生资产编号错误，此值+1，判断返回
                if MultipleCycles != 0:
                    # 当下一次循环开始走这边，不管对错，才能循环起来
                    wd.implicitly_wait(30)
                    if ReturnAfterRefresh != 0:
                        time.sleep(3)  # 如果资产正确，会加载一个分配成功，导致变慢，所以等待3秒，wait会查到元素，导致失败，所以只能是sleep
                        element = wd.find_elements_by_xpath(  # 如果正确的资产，循环走这个路劲点击
                            '//*[@id="app"]/section/section/main/div[2]/div/div[1]/div[1]/div[2]/div[3]/div')
                    else:
                        element = wd.find_elements_by_xpath(  # 当资产都正常的时候，进入到再次点击申请人路径不一样
                            # 为什么错误的能正确点击呢？因为错误了点击返回按钮原路返回，而正确的确是刷新的结果
                            '//*[@id="app"]/section/section/main/div[2]/div/div[1]/div[1]/div[2]/div[3]/div')
                    for i in element:
                        wd.implicitly_wait(30)
                        i.click()

                    # 二次循环内点击重置用户信息按钮
                    time.sleep(1)
                    element = wd.find_elements_by_xpath(
                        '//*[@id="AppropriationForm_applyUser"]/div/span[1]')
                    for y in element:
                        y.click()

                    # 当进入第二次循环，点击了重置按钮后，需要再次打开填入用户信息的按你牛
                    element = wd.find_elements_by_xpath(
                        '//*[@id="app"]/section/section/main/div[2]/div/div[1]/div[1]/div[2]/div[3]/div')
                    for i in element:
                        wd.implicitly_wait(10)
                        i.click()
                    # 第二次循环时插入输入
                    wd.implicitly_wait(10)
                    x.send_keys(tableLists)  # 将读取到的key循环的加入的申请人中，那么每添加一次，意味着要查询一次，分配一次资产

                # 第一次循环时插入用户
                if MultipleCycles == 0:
                    time.sleep(2)
                    x.send_keys(tableLists)  # 将读取到的key循环的加入的申请人中，那么每添加一次，意味着要查询一次，分配一次资产


                time.sleep(3)
                # 当用户搜索到了，就会出现TannaFilter--optionText，而且是唯一，如过没有，就会出现<p>
                try:
                    wd.find_element_by_css_selector('[class="TannaFilter--optionText"]')
                    element = wd.find_elements_by_xpath(
                         '/html/body/div[4]/div/div/div/div[2]/ul/li')  # 点击的绝对路径，相对路径的ID每次都会随机变化，无法使用
                    for y in element:
                        wd.implicitly_wait(3)  # 等待用户信息出现
                        y.click()  # 点击用户信息，进行搜索用户
                except Exception as f:
                    ExcelUserError += 1
                    wb = load_workbook(ExcelPath)
                    wb1 = wb.active
                    wb1.cell(1 + ExcelUserError, 5, tableLists + ' 此用户不存在，请检查邮箱名')  # 录入错误信息到Excel
                    wb.save(ExcelPath)
                    print('###################  ', tableLists, ' 此用户不存在，请检查邮箱名 ###################')
                    continue  # 用户不存在，进入下一次循环


                # 点击分配资产按钮
                wd.implicitly_wait(3)
                element = wd.find_elements_by_xpath(
                    '//*[@id="app"]/section/section/main/div[2]/div/div[2]/div/div/div/div/div/div[3]/div/div/table/tbody/tr[1]/td/div/div/button[2]')
                for y in element:
                    wd.implicitly_wait(30)
                    y.click()

                # 点击资产分配内的重置按钮
                wd.implicitly_wait(30)
                element = wd.find_elements_by_xpath(
                    '/html/body/div[5]/div/div[2]/div/div[2]/div[2]/div[2]/div[2]/div[2]/a')
                for y in element:
                    wd.implicitly_wait(30)
                    y.click()

                wd.implicitly_wait(30)
                element = wd.find_elements_by_xpath(  # 切换到按编号批量录入，原因：正常搜索无法正常点到分配按钮
                    '/html/body/div[5]/div/div[2]/div/div[2]/div[2]/div[2]/div[1]/label[2]')
                for y in element:
                    wd.implicitly_wait(30)
                    y.click()

                wd.implicitly_wait(10)
                element = wd.find_elements_by_xpath(  # 将邮箱对应的资产编号键入到搜索栏中
                    '/html/body/div[5]/div/div[2]/div/div[2]/div[2]/div[2]/div[2]/div/div/span/textarea')
                for y in element:
                    wd.implicitly_wait(5)  # 影响等待分配资产错误的返回错误速度
                    y.send_keys(tableList[tableLists], '\n')

                element = wd.find_elements_by_xpath(  # 勾选资产
                    '/html/body/div[5]/div/div[2]/div/div[2]/div[2]/div[2]/div[3]/div/div/div/div/div/div/div[2]/div/div/table/tbody/tr/td[1]/span'
                )
                if element:  # 判断是否找到了资产，没找到返回资产分配失败的信息
                    for y in element:
                        time.sleep(2)
                        y.click()

                else:
                    ExcelCountError += 1
                    wb = load_workbook(ExcelPath)
                    wb1 = wb.active
                    wb1.cell(1 + ExcelCountError, 4, tableList[tableLists] + '资产分配失败，请检查资产状态或者资产编号是否正确')  # 录入错误信息到Excel
                    wb.save(ExcelPath)

                    print('###################  ', tableList[tableLists], '资产分配失败，请检查资产状态或者资产编号是否正确 ###################')
                    element = wd.find_elements_by_xpath(
                        '//*[@id="rcDialogTitle0"]/div/i'
                    )
                    for y in element:
                        returnCount += 1  # 最开始的for循环，想以此进行下一次循环
                        y.click()  # 分配资产失败后成功点击返回
                    if returnCount != 0:  # 资产未找到进行下一次循环判断
                        MultipleCycles = 0
                        MultipleCycles += 1
                        continue

                wd.implicitly_wait(10)
                element = wd.find_elements_by_xpath(  # 点击分配资产分配按钮
                    '/html/body/div[5]/div/div[2]/div/div[2]/div[2]/div[3]/div[2]/button')
                for y in element:
                    wd.implicitly_wait(10)
                    y.click()

                wd.implicitly_wait(10)
                element = wd.find_elements_by_xpath(  # 输入异常信息
                    '/html/body/div[6]/div/div[2]/div/div[2]/div[2]/div/div[2]/div/form/div/div[2]/div/textarea')
                for y in element:
                    wd.implicitly_wait(10)
                    y.send_keys('入职')

                wd.implicitly_wait(30)
                element = wd.find_elements_by_xpath(  # 错误信息栏点击分配按钮
                    '/html/body/div[6]/div/div[2]/div/div[2]/div[3]/div/button[2]')
                for y in element:
                    wd.implicitly_wait(30)
                    y.click()
                    print(tableList[tableLists], '资产分配成功！！！')
                    MultipleCycles = 0
                    MultipleCycles += 1
                    ReturnAfterRefresh = 0
                    ReturnAfterRefresh += 1
                continue
        print()
        print('执行完毕，错误信息已经输出到用户信息表中！！！')
        print()
        print('10秒后自动关闭！！！')
        time.sleep(10)
    except Exception as f:
        print()
        print()
        print('请关闭此脚本打开的浏览器后  ！！！重新运行此脚本！！！，否则无法运行,脚本将在10秒后自动关闭！！！')
        print(f)
        time.sleep(10)
